import calendar
import json
import random
from datetime import date

import yaml
from openpyxl import load_workbook

# Cell references (adjust to match your template)
CELL_COMPANY = "B4"
CELL_NAME = "B5"
CELL_ROLE = "B6"
CELL_MONTH_YEAR = "F3"
DATE_START_ROW = 10
COL_DATE = "A"
COL_DAY = "B"
COL_DETAIL = "C"
COL_DAY_COUNT = "D"
CELL_PO_NUMBER = "F4"
CELL_QUOTATION_REFERENCE = "F5"


def load_config(path="config.yaml"):
    with open(path, "r") as f:
        return yaml.safe_load(f)


def get_workdays(year, month):
    """Return list of workday dates (Mon-Fri) for the given month."""
    days_in_month = calendar.monthrange(year, month)[1]
    workdays = []
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        if d.weekday() < 5:
            workdays.append(d)
    return workdays


def generate_descriptions_list(cfg):
    """Cycle through works list in original order."""
    workdays = get_workdays(cfg["year"], cfg["month"])
    works = cfg["works"]

    descriptions = {}
    for i, d in enumerate(workdays):
        descriptions[d.strftime("%Y-%m-%d")] = works[i % len(works)]

    return descriptions


def generate_descriptions_ai(cfg):
    """Use Claude to generate daily work descriptions."""
    import anthropic

    year = cfg["year"]
    month = cfg["month"]
    workdays = get_workdays(year, month)
    workday_strs = [d.strftime("%Y-%m-%d (%A)") for d in workdays]
    notes = cfg.get("notes", "")
    works = cfg.get("works", [])

    works_context = ""
    if works:
        works_context = f"""
Reference tasks (use these as inspiration, rephrase and vary them):
{chr(10).join(f'- {w}' for w in works)}
"""

    prompt = f"""Generate realistic daily work descriptions for a timesheet.

Role: {cfg['role']}
Project: {cfg['project']}
Month: {calendar.month_name[month]} {year}
Working days: {', '.join(workday_strs)}

{works_context}

Context/notes for this month:
{notes if notes.strip() else "General development work"}

Rules:
- One short description per workday (10-30 words)
- Vary the descriptions, don't repeat the same phrase
- Sound natural and professional
- Include realistic tasks: meetings, development, testing, code review, deployment, documentation, etc.
- Don't make every day sound exciting, some days are routine

Return ONLY a JSON object mapping date (YYYY-MM-DD) to description string.
Example: {{"2026-03-02": "Sprint planning meeting, started working on data pipeline refactoring"}}
"""

    client = anthropic.Anthropic(api_key=cfg["anthropic_api_key"])
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2000,
        messages=[{"role": "user", "content": prompt}],
    )

    text = response.content[0].text
    start = text.index("{")
    end = text.rindex("}") + 1
    return json.loads(text[start:end])


def generate_descriptions(cfg):
    """Route to the correct generation method based on mode."""
    mode = cfg.get("mode", "random")

    if mode == "ai":
        return generate_descriptions_ai(cfg)
    elif mode == "list":
        return generate_descriptions_list(cfg)
    else:
        raise ValueError(f"Unknown mode: {mode}. Use 'ai' or 'work list'.")


def fill_timesheet(cfg, descriptions):
    """Fill the timesheet Excel file."""
    year = cfg["year"]
    month = cfg["month"]

    wb = load_workbook(cfg["template_file"])
    ws = wb.active

    # --- Header fields ---
    ws[CELL_COMPANY] = cfg["company"]
    ws[CELL_NAME] = cfg["name"]
    ws[CELL_ROLE] = cfg["role"]
    ws[CELL_PO_NUMBER] = cfg.get("po_number", "")
    ws[CELL_QUOTATION_REFERENCE] = cfg.get("quotation_reference", "")

    month_abbr = calendar.month_abbr[month]
    year_short = str(year)[-2:]
    cell = ws[CELL_MONTH_YEAR]
    cell.value = f"{month_abbr}-{year_short}"
    cell.number_format = "@"  # force text format

    # --- Date rows ---
    days_in_month = calendar.monthrange(year, month)[1]

    for i in range(31):
        row = DATE_START_ROW + i
        if i < days_in_month:
            d = date(year, month, i + 1)
            date_key = d.strftime("%Y-%m-%d")

            ws[f"{COL_DATE}{row}"] = d.strftime("%-d %b %Y")
            ws[f"{COL_DAY}{row}"] = d.strftime("%a")

            if date_key in descriptions:
                ws[f"{COL_DETAIL}{row}"] = descriptions[date_key]
                ws[f"{COL_DAY_COUNT}{row}"] = 1
            else:
                ws[f"{COL_DETAIL}{row}"] = ""
                ws[f"{COL_DAY_COUNT}{row}"] = ""
        else:
            ws[f"{COL_DATE}{row}"] = ""
            ws[f"{COL_DAY}{row}"] = ""
            ws[f"{COL_DETAIL}{row}"] = ""
            ws[f"{COL_DAY_COUNT}{row}"] = ""

    wb.save(cfg["output_file"])


def main():
    cfg = load_config()
    mode = cfg.get("mode", "random")
    month_abbr = calendar.month_abbr[cfg["month"]]

    print(f"🔥 Timesheet: {cfg['name']} | {month_abbr} {cfg['year']}")
    print(f"{cfg['company']}")
    print(f"{cfg['role']} — {cfg['project']}")
    print(f"Mode: {mode}")
    print()

    print("Generating descriptions..." if mode == "ai" else "🎲 Picking from works list...")
    descriptions = generate_descriptions(cfg)
    print(f"Generated {len(descriptions)} workday descriptions")
    print()

    for date_key in sorted(descriptions.keys()):
        d = date.fromisoformat(date_key)
        day_name = d.strftime("%a")
        print(f"  {date_key} ({day_name}): {descriptions[date_key]}")
    print()

    fill_timesheet(cfg, descriptions)
    print(f"✅ Saved: {cfg['output_file']}")


if __name__ == "__main__":
    main()